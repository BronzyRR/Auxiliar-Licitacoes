from openpyxl import load_workbook
import xlrd
import xlsxwriter
import math
from ModuloGeral import *

# O objetivo desse método é retornar uma lista contendo os dados das cotações
def lerCotacoesPrefeituraAntiga(enderecoArquivo, nomePlanilha, linhaInicio, linhaFim):
    cotacoes = []

    planilhaGeral = xlrd.open_workbook(enderecoArquivo)
    planilhaCotacoes = planilhaGeral.sheet_by_name(nomePlanilha)

    for i in range(linhaInicio - 1, linhaFim):
        if (planilhaCotacoes.cell_value(rowx=i, colx=1) == "COTAÇÃO"):
            codigo = planilhaCotacoes.cell_value(rowx=i, colx=2)
            descricao = planilhaCotacoes.cell_value(rowx=i, colx=3)
            unidade = planilhaCotacoes.cell_value(rowx=i, colx=4)
            mediana = planilhaCotacoes.cell_value(rowx=i, colx=5)

            cotacoes.append([codigo, descricao, unidade, mediana])

    return cotacoes

# O objetivo desse método é, por meio de uma lista de cotações, criar uma planilha de maneira organizada para
# importar no OrçaFascio
def criarPlanilhaCotacaoOrcaFascio(cotacoes, anexoCodigoCotacao, enderecoSalvar):
    enderecoSalvar += "\[01] Insumos para importar.xlsx"

    workbook = xlsxwriter.Workbook(enderecoSalvar)
    worksheetCotacoes = workbook.add_worksheet("Cotações para importar")

    # criar os formatos
    formato_titulo = workbook.add_format(
        {'bold': True, 'font_size': 14, 'align': 'center', 'valign': 'vcenter', 'font_name': 'Arial',
         'bg_color': '#77c98d'})
    formato_item = workbook.add_format(
        {'bold': False, 'font_size': 12, 'align': 'left', 'valign': 'vcenter', 'font_name': 'Arial'})

    worksheetCotacoes.set_column(0, 0, 12.00)
    worksheetCotacoes.set_column(1, 1, 50.00)
    worksheetCotacoes.set_column(2, 2, 82.00)
    worksheetCotacoes.set_column(3, 3, 25.00)
    worksheetCotacoes.set_column(4, 4, 50.00)
    worksheetCotacoes.set_column(5, 5, 50.00)

    worksheetCotacoes.write(0, 0, "TIPO", formato_titulo)
    worksheetCotacoes.write(0, 1, "CÓDIGO", formato_titulo)
    worksheetCotacoes.write(0, 2, "DESCRIÇÃO", formato_titulo)
    worksheetCotacoes.write(0, 3, "UNIDADE", formato_titulo)
    worksheetCotacoes.write(0, 4, "PREÇO UNITÁRIO DESONERADO", formato_titulo)
    worksheetCotacoes.write(0, 5, "PREÇO UNITÁRIO NÃO DESONERADO", formato_titulo)

    contadorLinha = 1
    for item in cotacoes:
        worksheetCotacoes.write(contadorLinha, 0, 4, formato_item)
        worksheetCotacoes.write(contadorLinha, 1, "{} {}".format(item[0], anexoCodigoCotacao), formato_item)
        worksheetCotacoes.write(contadorLinha, 2, item[1], formato_item)
        worksheetCotacoes.write(contadorLinha, 3, item[2], formato_item)
        worksheetCotacoes.write(contadorLinha, 4, math.floor(float(item[3])*100)/100, formato_item)
        worksheetCotacoes.write(contadorLinha, 5, math.floor(float(item[3])*100)/100, formato_item)

        contadorLinha += 1

    workbook.close()


# O objetivo desse método é retornar uma lista contendo as composições "duplicadas", com sua respectiva posição no
# documento
def procurarComposicoesDuplicadas(enderecoArquivo, nomePlanilha, linhaInicio, linhaFim):
    composicoesDuplicadas = []

    planilhaGeral = xlrd.open_workbook(enderecoArquivo)
    planilhaComposicoes = planilhaGeral.sheet_by_name(nomePlanilha)

    # É uma variável booleana para dizer se o algoritmo está ou não dentro de uma composição
    flagEstouDentroDeUmaComposicao = False
    composicaoAtual = ""
    for i in range(linhaInicio - 1, linhaFim):

        if flagEstouDentroDeUmaComposicao:

            if planilhaComposicoes.cell_value(rowx=i, colx=1) == "":
                flagEstouDentroDeUmaComposicao = False

            elif planilhaComposicoes.cell_value(rowx=i, colx=2) == composicaoAtual:
                codigo = str(planilhaComposicoes.cell_value(rowx=i, colx=2))
                descricao = planilhaComposicoes.cell_value(rowx=i, colx=3)
                unidade = planilhaComposicoes.cell_value(rowx=i, colx=4)
                mediana = planilhaComposicoes.cell_value(rowx=i, colx=8)

                composicoesDuplicadas.append([codigo, descricao, unidade, mediana])

        else:
            if planilhaComposicoes.cell_value(rowx=i, colx=1) != "":
                flagEstouDentroDeUmaComposicao = True
                composicaoAtual = planilhaComposicoes.cell_value(rowx=i, colx=2)

    return composicoesDuplicadas


# O objetivo desse método é editar uma planilha de composição para importá-la no OrçaFascio
def criarPlanilhaComposicao(enderecoArquivo, nomePlanilha, linhaInicio, linhaFim, codigoComposicao,
                            codigoComposicaoInsumo, enderecoSalvar):

    enderecoSalvar += "\[02] Composições para importar.xlsx"

    workbook = xlsxwriter.Workbook(enderecoSalvar)
    worksheetComposicoes = workbook.add_worksheet("Composições para importar")

    # criar os formatos
    formato_titulo = workbook.add_format({'bold': True, 'font_size': 14, 'align': 'center', 'valign': 'vcenter',
                                          'font_name': 'Arial', 'bg_color': '#ff9985'})

    formato_composicao_titulo = workbook.add_format(
        {'bold': True, 'font_size': 8, 'align': 'left', 'valign': 'vcenter', 'font_name': 'Arial',
         'bg_color': '#ffe675'})

    formato_composicao_titulo_valores = workbook.add_format(
        {'bold': True, 'font_size': 8, 'align': 'left', 'valign': 'vcenter', 'font_name': 'Arial',
         'bg_color': '#949391'})

    formato_codigo_composicao = workbook.add_format(
        {'bold': False, 'font_size': 8, 'align': 'left', 'valign': 'vcenter', 'font_name': 'Arial',
         'bg_color': '#ffe675'})

    formato_item = workbook.add_format(
        {'bold': False, 'font_size': 8, 'align': 'left', 'valign': 'vcenter', 'font_name': 'Arial'})

    worksheetComposicoes.set_column(0, 0, 11.00)
    worksheetComposicoes.set_column(1, 1, 28.00)
    worksheetComposicoes.set_column(2, 2, 18.00)
    worksheetComposicoes.set_column(3, 3, 90.00)
    worksheetComposicoes.set_column(4, 4, 22.00)
    worksheetComposicoes.set_column(5, 5, 22.00)
    worksheetComposicoes.set_column(6, 6, 22.00)
    worksheetComposicoes.set_column(7, 7, 32.00)

    worksheetComposicoes.write(0, 0, "BASE", formato_titulo)
    worksheetComposicoes.write(0, 1, "TIPO", formato_titulo)
    worksheetComposicoes.write(0, 2, "CÓDIGO", formato_titulo)
    worksheetComposicoes.write(0, 3, "DESCRIÇÃO", formato_titulo)
    worksheetComposicoes.write(0, 4, "UNIDADE", formato_titulo)
    worksheetComposicoes.write(0, 5, "COEFICIENTE", formato_titulo)
    worksheetComposicoes.write(0, 6, "DESONERADO", formato_titulo)
    worksheetComposicoes.write(0, 7, "NÃO DESONERADO", formato_titulo)

    linhaWorksheet = 1

    planilhaGeral = xlrd.open_workbook(enderecoArquivo)
    planilhaComposicoes = planilhaGeral.sheet_by_name(nomePlanilha)

    # É uma variável booleana para dizer se o algoritmo está ou não dentro de uma composição
    flagEstouDentroDeUmaComposicao = False

    for i in range(linhaInicio - 1, linhaFim):

        if flagEstouDentroDeUmaComposicao:

            if planilhaComposicoes.cell_value(rowx=i, colx=1) == "":
                flagEstouDentroDeUmaComposicao = False
                worksheetComposicoes.write(linhaWorksheet, 0, "")

            else:
                fonte = planilhaComposicoes.cell_value(rowx=i, colx=1)
                codigo = str(planilhaComposicoes.cell_value(rowx=i, colx=2))
                descricao = planilhaComposicoes.cell_value(rowx=i, colx=3)
                unidade = planilhaComposicoes.cell_value(rowx=i, colx=4)
                coeficiente = planilhaComposicoes.cell_value(rowx=i, colx=7)
                desonerado = float(planilhaComposicoes.cell_value(rowx=i, colx=8))
                nao_desonerado = float(planilhaComposicoes.cell_value(rowx=i, colx=9))

                if(len(codigo)>2):

                    if(codigo[-1] == "0" and codigo[-2] == "."):
                        codigo = codigo[0:-2]

                print("CODIGO QUE ACHEI = {}".format(codigo))

                # arredondando
                desonerado = math.floor(desonerado * 100) / 100
                nao_desonerado = math.floor(nao_desonerado * 100) / 100

                tipo = ""

                if fonte=="SINAPI-I" or fonte=="COTAÇÃO":
                    tipo = "INSUMO"
                    if not (codigo.isnumeric()):
                        codigo += " {}".format(codigoComposicaoInsumo)
                else:
                    tipo = "COMPOSIÇÃO AUXILIAR"
                    if not (codigo.isnumeric()):
                        codigo += " {}".format(codigoComposicao)

                if fonte=="SINAPI-I" or fonte=="SINAPI":
                    fonte = "SINAPI"
                else:
                    fonte = "PRÓPRIO"

                worksheetComposicoes.write(linhaWorksheet, 0, fonte, formato_codigo_composicao)
                worksheetComposicoes.write(linhaWorksheet, 1, tipo, formato_codigo_composicao)
                worksheetComposicoes.write(linhaWorksheet, 2, codigo, formato_codigo_composicao)
                worksheetComposicoes.write(linhaWorksheet, 3, descricao, formato_item)
                worksheetComposicoes.write(linhaWorksheet, 4, unidade, formato_item)
                worksheetComposicoes.write(linhaWorksheet, 5, coeficiente, formato_item)
                worksheetComposicoes.write(linhaWorksheet, 6, desonerado, formato_item)
                worksheetComposicoes.write(linhaWorksheet, 7, nao_desonerado, formato_item)

        else:
            #encontrei uma nova composição
            if planilhaComposicoes.cell_value(rowx=i, colx=1) != "":
                flagEstouDentroDeUmaComposicao = True

                codigo = str(planilhaComposicoes.cell_value(rowx=i, colx=2))
                descricao = planilhaComposicoes.cell_value(rowx=i, colx=3)
                unidade = planilhaComposicoes.cell_value(rowx=i, colx=4)
                coeficiente = ""
                desonerado = float(planilhaComposicoes.cell_value(rowx=i, colx=8))
                nao_desonerado = float(planilhaComposicoes.cell_value(rowx=i, colx=9))


                if (codigo[-1] == "0" and codigo[-2] == "."):
                    codigo = codigo[0:-2]

                print("CODIGO NOVA COMPOSIÇÃO = {}".format(codigo))

                #arredondando
                desonerado = math.floor(desonerado*100)/100
                nao_desonerado = math.floor(nao_desonerado*100)/100

                worksheetComposicoes.write(linhaWorksheet, 0, "PRÓPRIO", formato_composicao_titulo)
                worksheetComposicoes.write(linhaWorksheet, 1, "COMPOSIÇÃO", formato_composicao_titulo)
                worksheetComposicoes.write(linhaWorksheet, 2, "{} {}".format(codigo, codigoComposicao), formato_composicao_titulo)
                worksheetComposicoes.write(linhaWorksheet, 3, descricao, formato_composicao_titulo)
                worksheetComposicoes.write(linhaWorksheet, 4, unidade, formato_composicao_titulo)
                worksheetComposicoes.write(linhaWorksheet, 5, coeficiente, formato_composicao_titulo_valores)
                worksheetComposicoes.write(linhaWorksheet, 6, desonerado, formato_composicao_titulo_valores)
                worksheetComposicoes.write(linhaWorksheet, 7, nao_desonerado, formato_composicao_titulo_valores)

        linhaWorksheet += 1

    workbook.close()

# O objetivo deste métido é editar uma planilha de orçamento para importá-la no OrçaFascio
def criarPlanilhaOrçamento(enderecoArquivo, nomePlanilha, linhaInicio, linhaFim, codigoComposicao,
                           enderecoSalvar):

    enderecoSalvar += "\[03] Orçamento para importar.xlsx"

    workbook = xlsxwriter.Workbook(enderecoSalvar)
    worksheetOrcamento = workbook.add_worksheet("Orçamentos para importar")

    # criar os formatos
    formato_titulo_geral = workbook.add_format({'bold': True, 'font_size': 14, 'align': 'center', 'valign': 'vcenter',
                                          'font_name': 'Arial', 'bg_color': '#ff9985'})

    formato_titulo_orçamento = workbook.add_format(
        {'bold': True, 'font_size': 11, 'align': 'left', 'valign': 'vcenter', 'font_name': 'Arial',
         'bg_color': '#a3a3a2'})

    formato_item = workbook.add_format(
        {'bold': False, 'font_size': 9, 'align': 'center', 'valign': 'vcenter', 'font_name': 'Arial'})

    formato_descricao = workbook.add_format(
        {'bold': False, 'font_size': 9, 'align': 'left', 'valign': 'vcenter', 'font_name': 'Arial'})

    worksheetOrcamento.set_column(0, 0, 12.00)
    worksheetOrcamento.set_column(1, 1, 15.00)
    worksheetOrcamento.set_column(2, 2, 22.00)
    worksheetOrcamento.set_column(3, 3, 90.00)
    worksheetOrcamento.set_column(4, 4, 22.00)

    worksheetOrcamento.write(0, 0, "ITEM", formato_titulo_geral)
    worksheetOrcamento.write(0, 1, "FONTE", formato_titulo_geral)
    worksheetOrcamento.write(0, 2, "CÓDIGO", formato_titulo_geral)
    worksheetOrcamento.write(0, 3, "DESCRIÇÃO", formato_titulo_geral)
    worksheetOrcamento.write(0, 4, "QUANTIDADE", formato_titulo_geral)


    #planilha para ler os dados
    planilhaGeral = xlrd.open_workbook(enderecoArquivo)
    planilhaOrcamento = planilhaGeral.sheet_by_name(nomePlanilha)

    #para verificar em qual item estamos
    meta = 0
    nivel_01 = 0
    nivel_02 = 0
    nivel_03 = 0
    nivel_04 = 0
    itemAtual = 0

    linhaWorksheet = 1
    for i in range(linhaInicio - 1, linhaFim):
        nivel = planilhaOrcamento.cell_value(rowx=i, colx=12)
        descricao = planilhaOrcamento.cell_value(rowx=i, colx=17)

        if nivel.upper() == "META":
            meta += 1
            itemAtual = 0
        elif nivel.upper() == "NÍVEL 1":
            nivel_01 += 1
            nivel_02 = 0
            nivel_03 = 0
            nivel_04 = 0
            itemAtual = 0
        elif nivel.upper() == "NÍVEL 2":
            nivel_02 += 1
            nivel_03 = 0
            nivel_04 = 0
            itemAtual = 0
        elif nivel.upper() == "NÍVEL 3":
            nivel_03 += 1
            nivel_04 = 0
            itemAtual = 0
        elif nivel.upper() == "NÍVEL 4":
            nivel_04 += 1
            itemAtual = 0

        if nivel.upper() == "SERVIÇO":
            itemAtual+=1
            codigo = str(planilhaOrcamento.cell_value(rowx=i, colx=16))
            quantidade = planilhaOrcamento.cell_value(rowx=i, colx=19)
            fonte = ""

            if codigo.isnumeric():
                fonte = "SINAPI"
            else:
                fonte = "PRÓPRIO"
                codigo += " {}".format(codigoComposicao)

            worksheetOrcamento.write(linhaWorksheet, 0, retornarStringItem(meta, nivel_01, nivel_02, nivel_03, nivel_04, itemAtual), formato_item)
            worksheetOrcamento.write(linhaWorksheet, 1, fonte, formato_item)
            worksheetOrcamento.write(linhaWorksheet, 2, codigo, formato_item)
            worksheetOrcamento.write(linhaWorksheet, 3, descricao, formato_descricao)
            worksheetOrcamento.write(linhaWorksheet, 4, quantidade, formato_item)

        else:
            worksheetOrcamento.write(linhaWorksheet, 0, retornarStringItem(meta, nivel_01, nivel_02, nivel_03, nivel_04, itemAtual), formato_titulo_orçamento)
            worksheetOrcamento.write(linhaWorksheet, 1, "", formato_titulo_orçamento)
            worksheetOrcamento.write(linhaWorksheet, 2, "", formato_titulo_orçamento)
            worksheetOrcamento.write(linhaWorksheet, 3, descricao, formato_titulo_orçamento)
            worksheetOrcamento.write(linhaWorksheet, 4, "", formato_titulo_orçamento)

        linhaWorksheet += 1

    workbook.close()



print("*************** 1° PASSO: INFORMAR OS ENDEREÇOS ***************")
enderecoArquivo = input("Informe o caminho do arquivo de referênica: ")
enderecoArquivo = enderecoArquivo[1:(len(enderecoArquivo) - 1)]

enderecoArquivoOrcamento = input("Informe o caminho do arquivo de orçamento: ")
enderecoArquivoOrcamento = enderecoArquivoOrcamento[1:(len(enderecoArquivoOrcamento) - 1)]

enderecoSalvar = input("Informe a pasta em que deseja salvar os resultados: ")
enderecoSalvar = enderecoSalvar[1:(len(enderecoSalvar) - 1)]

print("\n\n*************** 2° PASSO: COTAÇÕES ***************")
nomePlanilhaCotacao = input("Informe o nome da planilha de cotações: ")
linhaInicioCotacao = int(input("Informe a linha de início das cotações: "))
linhaFimCotacao = int(input("Informe a linha de fim das cotações: "))

cotacoes = lerCotacoesPrefeituraAntiga(enderecoArquivo, nomePlanilhaCotacao, linhaInicioCotacao, linhaFimCotacao)

print("\n\n*************** 3° PASSO: PROCURAR COMPOSIÇÕES DUPLICADAS ***************")
nomePlanilhaComposicoes = input("Informe o nome da planilha de composições: ")
linhaInicioComposicao = int(input("Informe a linha de início das composições: "))
linhaFimComposicao = int(input("Informe a linha de fim das composições: "))

composicoesDuplicadas = procurarComposicoesDuplicadas(enderecoArquivo, nomePlanilhaComposicoes, linhaInicioComposicao, linhaFimComposicao)

print("\n\n*************** 4° PASSO: ORÇAMENTO ***************")
nomePlanilhaOrcamento = input("Informe o nome da planilha de orçamento: ")
linhaInicioOrcamento = int(input("Informe a linha de início do orçamento (META): "))
linhaFimOrcamento = int(input("Informe a linha de fim do orçamento: "))

print("\n\n*************** 5° PASSO: CRIAR PLANILHAS ***************")
concorrencia = input("Informe o nome da concorrência (Ex: CP009): ")
codigoComposicao = concorrencia
codigoComposicaoInsumo = concorrencia + "_INSUMO"
cotacoes += composicoesDuplicadas

criarPlanilhaCotacaoOrcaFascio(cotacoes, codigoComposicaoInsumo, enderecoSalvar)

criarPlanilhaComposicao(enderecoArquivo, nomePlanilhaComposicoes, linhaInicioComposicao, linhaFimComposicao,
                        codigoComposicao, codigoComposicaoInsumo, enderecoSalvar)

criarPlanilhaOrçamento(enderecoArquivoOrcamento, nomePlanilhaOrcamento, linhaInicioOrcamento, linhaFimOrcamento,
                       codigoComposicao, enderecoSalvar)

print("\n\n*************** PLANILHAS CRIADAS COM SUCESSO >-< ***************")







